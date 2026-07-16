"""
Tests for customizable visible fields and tire renewal features.

Covers:
- Model: visible_fields property, tires_need_renewal column
- Migration: ALTER TABLE adds new columns
- Routes: settings save persists visible_fields_json; create/edit
          handle tires_need_renewal; list filters by renewal
- Templates: settings page shows field checkboxes; form shows renewal;
             list shows renewal badge and filter
"""
import json
import secrets

from sqlalchemy import create_engine, inspect, text, event

from tsm.models import Settings, WheelSet


# ── Helpers ────────────────────────────────────────────────────────────────

def _csrf(client):
    with client.session_transaction() as sess:
        tok = secrets.token_urlsafe(16)
        sess["_csrf_token"] = tok
    return tok


# ── Model tests ────────────────────────────────────────────────────────────

class TestVisibleFieldsModel:
    def test_visible_fields_json_column_exists(self):
        cols = {c.key for c in Settings.__table__.columns}
        assert "visible_fields_json" in cols

    def test_visible_fields_default_empty(self, db_session):
        s = Settings(backup_interval_minutes=60, backup_copies=10)
        db_session.add(s)
        db_session.commit()
        db_session.expire(s)
        assert s.visible_fields == []

    def test_visible_fields_setter(self, db_session):
        s = Settings(backup_interval_minutes=60, backup_copies=10)
        db_session.add(s)
        db_session.commit()
        s.visible_fields = ["tire_manufacturer", "tire_size", "season"]
        db_session.commit()
        db_session.expire(s)
        assert set(s.visible_fields) == {"tire_manufacturer", "tire_size", "season"}

    def test_visible_fields_setter_filters_invalid(self, db_session):
        s = Settings(backup_interval_minutes=60, backup_copies=10)
        db_session.add(s)
        db_session.commit()
        s.visible_fields = ["tire_manufacturer", "invalid_field", "tire_size"]
        db_session.commit()
        db_session.expire(s)
        assert set(s.visible_fields) == {"tire_manufacturer", "tire_size"}

    def test_is_field_visible_when_tire_details_enabled(self, db_session):
        s = Settings(backup_interval_minutes=60, backup_copies=10,
                     enable_tire_details=True)
        db_session.add(s)
        db_session.commit()
        # All fields visible when enable_tire_details is True
        assert s.is_field_visible("season") is True
        assert s.is_field_visible("tire_manufacturer") is True

    def test_is_field_visible_with_visible_fields(self, db_session):
        s = Settings(backup_interval_minutes=60, backup_copies=10,
                     enable_tire_details=False)
        db_session.add(s)
        db_session.commit()
        s.visible_fields = ["tire_size", "season"]
        db_session.commit()
        assert s.is_field_visible("tire_size") is True
        assert s.is_field_visible("season") is True
        assert s.is_field_visible("tire_manufacturer") is False


class TestTiresNeedRenewalModel:
    def test_column_exists(self):
        cols = {c.key for c in WheelSet.__table__.columns}
        assert "tires_need_renewal" in cols

    def test_default_false(self, db_session):
        ws = WheelSet(
            customer_name="Test", license_plate="B-XX 1",
            car_type="Golf", storage_position="C1ROM",
        )
        db_session.add(ws)
        db_session.commit()
        db_session.expire(ws)
        assert ws.tires_need_renewal is False

    def test_can_set_true(self, db_session):
        ws = WheelSet(
            customer_name="Test", license_plate="B-XX 2",
            car_type="Golf", storage_position="C1LOM",
            tires_need_renewal=True,
        )
        db_session.add(ws)
        db_session.commit()
        db_session.expire(ws)
        assert ws.tires_need_renewal is True


# ── Migration tests ────────────────────────────────────────────────────────

class TestMigrationNewColumns:
    """Verify _migrate() adds new columns to an old database."""

    def _make_old_engine(self):
        eng = create_engine(
            "sqlite:///:memory:",
            echo=False,
            future=True,
            connect_args={"check_same_thread": False},
        )

        @event.listens_for(eng, "connect")
        def _pragma(dbapi_conn, _rec):
            cur = dbapi_conn.cursor()
            cur.execute("PRAGMA journal_mode=WAL;")
            cur.close()

        with eng.begin() as conn:
            # Old settings table (without visible_fields_json)
            conn.execute(text("""
                CREATE TABLE settings (
                    id INTEGER PRIMARY KEY,
                    backup_interval_minutes INTEGER NOT NULL DEFAULT 60,
                    backup_copies INTEGER NOT NULL DEFAULT 10,
                    dark_mode BOOLEAN NOT NULL DEFAULT 0,
                    auto_update BOOLEAN NOT NULL DEFAULT 1,
                    language VARCHAR(10) NOT NULL DEFAULT 'de',
                    custom_positions_json TEXT,
                    enable_tire_details BOOLEAN NOT NULL DEFAULT 0,
                    enable_seasonal_tracking BOOLEAN NOT NULL DEFAULT 0,
                    updated_at DATETIME
                )
            """))
            # Old wheel_sets table (without tires_need_renewal)
            conn.execute(text("""
                CREATE TABLE wheel_sets (
                    id INTEGER PRIMARY KEY,
                    customer_name VARCHAR(200) NOT NULL,
                    license_plate VARCHAR(20) NOT NULL,
                    car_type VARCHAR(100) NOT NULL,
                    storage_position VARCHAR(20) NOT NULL UNIQUE,
                    note TEXT,
                    tire_manufacturer VARCHAR(100),
                    tire_size VARCHAR(50),
                    tire_age VARCHAR(20),
                    season VARCHAR(20),
                    rim_type VARCHAR(20),
                    exchange_note TEXT,
                    created_at DATETIME,
                    updated_at DATETIME
                )
            """))
            # audit_log needed for _migrate to pass
            conn.execute(text("""
                CREATE TABLE audit_log (
                    id INTEGER PRIMARY KEY,
                    action VARCHAR(50) NOT NULL,
                    wheelset_id INTEGER,
                    details TEXT,
                    created_at DATETIME
                )
            """))
        return eng

    def test_visible_fields_json_added(self, monkeypatch):
        eng = self._make_old_engine()
        import tsm.db as db_mod
        monkeypatch.setattr(db_mod, "engine", eng)
        db_mod._migrate()
        insp = inspect(eng)
        cols = {c["name"] for c in insp.get_columns("settings")}
        assert "visible_fields_json" in cols

    def test_tires_need_renewal_added(self, monkeypatch):
        eng = self._make_old_engine()
        import tsm.db as db_mod
        monkeypatch.setattr(db_mod, "engine", eng)
        db_mod._migrate()
        insp = inspect(eng)
        cols = {c["name"] for c in insp.get_columns("wheel_sets")}
        assert "tires_need_renewal" in cols


# ── Route tests ────────────────────────────────────────────────────────────

class TestSettingsVisibleFieldsRoute:
    def test_save_visible_fields(self, client, seed_settings, db_session):
        tok = _csrf(client)
        resp = client.post("/settings", data={
            "_csrf_token": tok,
            "_visible_fields_submitted": "1",
            "backup_interval_minutes": "60",
            "backup_copies": "10",
            "dark_mode": "0",
            "auto_update": "0",
            "language": "de",
            "enable_tire_details": "0",
            "enable_seasonal_tracking": "0",
            "visible_fields": ["tire_manufacturer", "tire_size", "season"],
        }, follow_redirects=True)
        assert resp.status_code == 200
        s = db_session.query(Settings).first()
        assert set(s.visible_fields) == {"tire_manufacturer", "tire_size", "season"}

    def test_visible_fields_checkboxes_shown_when_tire_details_off(
            self, client, seed_settings):
        resp = client.get("/settings")
        assert resp.status_code == 200
        assert b"vf_tire_manufacturer" in resp.data
        assert b"vf_tire_size" in resp.data
        # 'note' is NOT in OPTIONAL_FIELDS — no checkbox for it
        assert b"vf_note" not in resp.data

    def test_visible_fields_hidden_when_tire_details_on(
            self, client, seed_settings, db_session):
        seed_settings.enable_tire_details = True
        db_session.commit()
        resp = client.get("/settings")
        assert resp.status_code == 200
        assert b"vf_note" not in resp.data


class TestTireRenewalRoutes:
    def test_create_with_renewal_flag(self, client, seed_settings, db_session):
        tok = _csrf(client)
        resp = client.post("/wheelsets/new", data={
            "_csrf_token": tok,
            "customer_name": "Anna Müller",
            "license_plate": "B-AM 1234",
            "car_type": "BMW 3er",
            "storage_position": "C1ROM",
            "tires_need_renewal": "1",
        }, follow_redirects=True)
        assert resp.status_code == 200
        ws = db_session.query(WheelSet).filter_by(
            license_plate="B-AM 1234").first()
        assert ws is not None
        assert ws.tires_need_renewal is True

    def test_create_without_renewal_flag(self, client, seed_settings, db_session):
        tok = _csrf(client)
        resp = client.post("/wheelsets/new", data={
            "_csrf_token": tok,
            "customer_name": "Bob Schmidt",
            "license_plate": "M-BS 5678",
            "car_type": "Audi A4",
            "storage_position": "C1ROM",
        }, follow_redirects=True)
        assert resp.status_code == 200
        ws = db_session.query(WheelSet).filter_by(
            license_plate="M-BS 5678").first()
        assert ws is not None
        assert ws.tires_need_renewal is False

    def test_edit_sets_renewal_flag(self, client, seed_wheelset,
                                    seed_settings, db_session):
        tok = _csrf(client)
        wid = seed_wheelset.id
        resp = client.post(f"/wheelsets/{wid}/edit", data={
            "_csrf_token": tok,
            "customer_name": seed_wheelset.customer_name,
            "license_plate": seed_wheelset.license_plate,
            "car_type": seed_wheelset.car_type,
            "storage_position": seed_wheelset.storage_position,
            "tires_need_renewal": "1",
        }, follow_redirects=True)
        assert resp.status_code == 200
        ws = db_session.get(WheelSet, wid)
        assert ws.tires_need_renewal is True

    def test_filter_renewal_in_list(self, client, seed_settings, db_session):
        # Create two wheelsets — one with renewal, one without
        ws1 = WheelSet(
            customer_name="Renewal", license_plate="B-RN 1",
            car_type="Golf", storage_position="C1ROM",
            tires_need_renewal=True,
        )
        ws2 = WheelSet(
            customer_name="NoRenewal", license_plate="B-NR 2",
            car_type="Golf", storage_position="C1LOM",
            tires_need_renewal=False,
        )
        db_session.add_all([ws1, ws2])
        db_session.commit()

        resp = client.get("/wheelsets?filter_renewal=1")
        assert resp.status_code == 200
        assert b"Renewal" in resp.data
        assert b"NoRenewal" not in resp.data

    def test_renewal_badge_shown_in_list(self, client, seed_settings, db_session):
        ws = WheelSet(
            customer_name="BadTires", license_plate="B-BT 3",
            car_type="Polo", storage_position="C1ROM",
            tires_need_renewal=True,
        )
        db_session.add(ws)
        db_session.commit()

        resp = client.get("/wheelsets")
        assert resp.status_code == 200
        assert b"renewal-badge" in resp.data

    def test_renewal_checkbox_in_form(self, client, seed_settings):
        resp = client.get("/wheelsets/new")
        assert resp.status_code == 200
        assert b"tiresNeedRenewal" in resp.data


# ── Per-field visibility: routes persist individually enabled fields ────────

class TestPerFieldPersistence:
    """When enable_tire_details is OFF but a field is in visible_fields,
    the create/edit routes must still persist that field, and must ignore
    fields that are NOT visible (mass-assignment protection)."""

    def _enable_fields(self, db_session, seed_settings, fields):
        seed_settings.enable_tire_details = False
        seed_settings.visible_fields = fields
        db_session.commit()

    def test_create_persists_visible_field(
            self, client, seed_settings, db_session):
        self._enable_fields(db_session, seed_settings, ["season"])
        tok = _csrf(client)
        resp = client.post("/wheelsets/new", data={
            "_csrf_token": tok,
            "customer_name": "Clara",
            "license_plate": "B-CL 100",
            "car_type": "Golf",
            "storage_position": "C1ROM",
            "season": "winter",
        }, follow_redirects=True)
        assert resp.status_code == 200
        ws = db_session.query(WheelSet).filter_by(
            license_plate="B-CL 100").first()
        assert ws is not None
        assert ws.season == "winter"

    def test_create_ignores_non_visible_field(
            self, client, seed_settings, db_session):
        """A field NOT in visible_fields must be ignored even if posted."""
        self._enable_fields(db_session, seed_settings, ["season"])
        tok = _csrf(client)
        resp = client.post("/wheelsets/new", data={
            "_csrf_token": tok,
            "customer_name": "Dora",
            "license_plate": "B-DO 200",
            "car_type": "Golf",
            "storage_position": "C1ROM",
            "season": "winter",
            # tire_manufacturer is NOT visible — must be ignored
            "tire_manufacturer": "Injected",
        }, follow_redirects=True)
        assert resp.status_code == 200
        ws = db_session.query(WheelSet).filter_by(
            license_plate="B-DO 200").first()
        assert ws is not None
        assert ws.season == "winter"
        assert ws.tire_manufacturer is None

    def test_edit_ignores_non_visible_field(
            self, client, seed_wheelset, seed_settings, db_session):
        self._enable_fields(db_session, seed_settings, ["tire_size"])
        wid = seed_wheelset.id
        tok = _csrf(client)
        resp = client.post(f"/wheelsets/{wid}/edit", data={
            "_csrf_token": tok,
            "customer_name": seed_wheelset.customer_name,
            "license_plate": seed_wheelset.license_plate,
            "car_type": seed_wheelset.car_type,
            "storage_position": seed_wheelset.storage_position,
            "tire_size": "225/45 R17",
            # rim_type NOT visible — must be ignored
            "rim_type": "alu",
        }, follow_redirects=True)
        assert resp.status_code == 200
        ws = db_session.get(WheelSet, wid)
        assert ws.tire_size == "225/45 R17"
        assert ws.rim_type is None

    def test_all_fields_saved_when_tire_details_on(
            self, client, seed_settings, db_session):
        seed_settings.enable_tire_details = True
        db_session.commit()
        tok = _csrf(client)
        resp = client.post("/wheelsets/new", data={
            "_csrf_token": tok,
            "customer_name": "Emil",
            "license_plate": "B-EM 300",
            "car_type": "Golf",
            "storage_position": "C1ROM",
            "tire_manufacturer": "Conti",
            "tire_size": "205/55 R16",
            "season": "sommer",
            "rim_type": "alu",
        }, follow_redirects=True)
        assert resp.status_code == 200
        ws = db_session.query(WheelSet).filter_by(
            license_plate="B-EM 300").first()
        assert ws.tire_manufacturer == "Conti"
        assert ws.tire_size == "205/55 R16"
        assert ws.season == "sommer"
        assert ws.rim_type == "alu"


# ── Per-field visibility: form rendering ────────────────────────────────────

class TestPerFieldFormRendering:
    def test_form_shows_only_visible_field(
            self, client, seed_settings, db_session):
        seed_settings.enable_tire_details = False
        seed_settings.visible_fields = ["season"]
        db_session.commit()
        resp = client.get("/wheelsets/new")
        assert resp.status_code == 200
        # season field visible
        assert b'name="season"' in resp.data
        # tire_manufacturer NOT visible
        assert b'name="tire_manufacturer"' not in resp.data

    def test_form_hides_all_optional_when_none_visible(
            self, client, seed_settings, db_session):
        seed_settings.enable_tire_details = False
        seed_settings.visible_fields = []
        db_session.commit()
        resp = client.get("/wheelsets/new")
        assert resp.status_code == 200
        assert b'name="season"' not in resp.data
        assert b'name="tire_manufacturer"' not in resp.data
        assert b'name="rim_type"' not in resp.data

    def test_form_shows_all_when_tire_details_on(
            self, client, seed_settings, db_session):
        seed_settings.enable_tire_details = True
        db_session.commit()
        resp = client.get("/wheelsets/new")
        assert resp.status_code == 200
        assert b'name="tire_manufacturer"' in resp.data
        assert b'name="season"' in resp.data
        assert b'name="rim_type"' in resp.data


# ── Per-field visibility: list rendering & filters ──────────────────────────

class TestPerFieldListRendering:
    def test_season_badge_shown_when_season_visible(
            self, client, seed_settings, db_session):
        seed_settings.enable_tire_details = False
        seed_settings.visible_fields = ["season"]
        db_session.commit()
        ws = WheelSet(
            customer_name="Frida", license_plate="B-FR 400",
            car_type="Golf", storage_position="C1ROM", season="winter",
        )
        db_session.add(ws)
        db_session.commit()
        resp = client.get("/wheelsets")
        assert resp.status_code == 200
        assert b"season-badge" in resp.data

    def test_season_badge_hidden_when_season_not_visible(
            self, client, seed_settings, db_session):
        seed_settings.enable_tire_details = False
        seed_settings.visible_fields = []
        db_session.commit()
        ws = WheelSet(
            customer_name="Gustav", license_plate="B-GU 500",
            car_type="Golf", storage_position="C1ROM", season="winter",
        )
        db_session.add(ws)
        db_session.commit()
        resp = client.get("/wheelsets")
        assert resp.status_code == 200
        assert b"season-badge" not in resp.data

    def test_season_filter_shown_when_visible(
            self, client, seed_settings, db_session):
        seed_settings.enable_tire_details = False
        seed_settings.visible_fields = ["season"]
        db_session.commit()
        resp = client.get("/wheelsets")
        assert resp.status_code == 200
        assert b'name="filter_season"' in resp.data

    def test_season_filter_hidden_when_not_visible(
            self, client, seed_settings, db_session):
        seed_settings.enable_tire_details = False
        seed_settings.visible_fields = []
        db_session.commit()
        resp = client.get("/wheelsets")
        assert resp.status_code == 200
        assert b'name="filter_season"' not in resp.data


# ── Sentinel guard: other settings forms must not clear visible_fields ──────

class TestVisibleFieldsSentinel:
    def test_other_form_does_not_clear_visible_fields(
            self, client, seed_settings, db_session):
        # Pre-set visible fields
        seed_settings.visible_fields = ["tire_size", "season"]
        db_session.commit()

        # Simulate the dark-mode toggle form (no visible_fields inputs,
        # no sentinel) — visible_fields must be preserved.
        tok = _csrf(client)
        resp = client.post("/settings", data={
            "_csrf_token": tok,
            "backup_interval_minutes": "60",
            "backup_copies": "10",
            "dark_mode": "1",
            "auto_update": "0",
            "language": "de",
            "enable_tire_details": "0",
            "enable_seasonal_tracking": "0",
        }, follow_redirects=True)
        assert resp.status_code == 200
        s = db_session.query(Settings).first()
        assert set(s.visible_fields) == {"tire_size", "season"}

    def test_sentinel_form_can_clear_visible_fields(
            self, client, seed_settings, db_session):
        seed_settings.visible_fields = ["tire_size", "season"]
        db_session.commit()

        tok = _csrf(client)
        # Submit the visible-fields form with sentinel but no checkboxes
        resp = client.post("/settings", data={
            "_csrf_token": tok,
            "_visible_fields_submitted": "1",
            "backup_interval_minutes": "60",
            "backup_copies": "10",
            "dark_mode": "0",
            "auto_update": "0",
            "language": "de",
            "enable_tire_details": "0",
            "enable_seasonal_tracking": "0",
        }, follow_redirects=True)
        assert resp.status_code == 200
        s = db_session.query(Settings).first()
        assert s.visible_fields == []


# ── Export tests: CSV & XLSX include new columns ───────────────────────────

class TestCsvExportNewFields:
    def test_csv_header_includes_tire_fields(self, db_session, db_engine,
                                             monkeypatch):
        import os, tempfile
        import tsm.backup_manager as bm_mod
        monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.csv")
            from tsm.backup_manager import export_csv_snapshot
            export_csv_snapshot(target)
            with open(target, encoding="utf-8-sig") as f:
                header = f.readline()
            assert "tire_manufacturer" in header
            assert "tires_need_renewal" in header
            assert "season" in header

    def test_csv_data_includes_tire_values(self, db_session, db_engine,
                                           monkeypatch):
        import os, tempfile
        import tsm.backup_manager as bm_mod
        monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
        ws = WheelSet(
            customer_name="Export", license_plate="B-EX 1",
            car_type="Golf", storage_position="C1ROM",
            tire_manufacturer="Conti", season="winter",
            tires_need_renewal=True,
        )
        db_session.add(ws)
        db_session.commit()
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.csv")
            from tsm.backup_manager import export_csv_snapshot
            export_csv_snapshot(target)
            with open(target, encoding="utf-8-sig") as f:
                lines = f.readlines()
            data = lines[1]
            assert "Conti" in data
            assert "winter" in data
            # tires_need_renewal exported as "1"
            assert ";1;" in data


class TestXlsxExportNewFields:
    def test_xlsx_header_includes_tire_columns(self, db_session, db_engine,
                                                seed_wheelset, monkeypatch):
        import os, tempfile
        import tsm.backup_manager as bm_mod
        monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.xlsx")
            from tsm.backup_manager import export_xlsx_snapshot
            export_xlsx_snapshot(target)
            from openpyxl import load_workbook
            wb = load_workbook(target)
            sheet = wb.active
            # Find column header row — scan for "Hersteller"
            found = False
            for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
                if row and "Hersteller" in (row or []):
                    found = True
                    break
            assert found, "XLSX should have 'Hersteller' column header"

    def test_xlsx_renewal_marker(self, db_session, db_engine, monkeypatch):
        import os, tempfile
        import tsm.backup_manager as bm_mod
        monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
        ws = WheelSet(
            customer_name="XlsxTest", license_plate="B-XL 1",
            car_type="Golf", storage_position="C1ROM",
            tires_need_renewal=True,
        )
        db_session.add(ws)
        db_session.commit()
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.xlsx")
            from tsm.backup_manager import export_xlsx_snapshot
            export_xlsx_snapshot(target)
            from openpyxl import load_workbook
            wb = load_workbook(target)
            sheet = wb.active
            # Find the ⚠ marker in the sheet
            found = False
            for row in sheet.iter_rows(values_only=True):
                if row and "\u26a0" in (str(c) for c in row if c):
                    found = True
                    break
            assert found, "XLSX should show ⚠ for renewal-flagged wheelsets"


# ── Inventory print: renewal column ───────────────────────────────────────

class TestInventoryPrintRenewal:
    def test_renewal_marker_in_print(self, client, seed_settings, db_session):
        ws = WheelSet(
            customer_name="PrintTest", license_plate="B-PT 1",
            car_type="Golf", storage_position="C1ROM",
            tires_need_renewal=True,
        )
        db_session.add(ws)
        db_session.commit()
        resp = client.get("/backups/inventory")
        assert resp.status_code == 200
        # The ⚠ warning sign should appear for renewal rows
        assert "⚠" in resp.data.decode("utf-8")

    def test_no_renewal_marker_when_not_flagged(self, client, seed_settings,
                                                 db_session):
        ws = WheelSet(
            customer_name="NoRenewalPrint", license_plate="B-NP 2",
            car_type="Golf", storage_position="C1ROM",
            tires_need_renewal=False,
        )
        db_session.add(ws)
        db_session.commit()
        resp = client.get("/backups/inventory")
        assert resp.status_code == 200
        assert "⚠" not in resp.data.decode("utf-8")

    def test_renewal_column_header_in_print(self, client, seed_settings,
                                             db_session):
        ws = WheelSet(
            customer_name="HeaderTest", license_plate="B-HT 3",
            car_type="Golf", storage_position="C1ROM",
        )
        db_session.add(ws)
        db_session.commit()
        resp = client.get("/backups/inventory")
        assert resp.status_code == 200
        # Column header uses &#x26A0; (⚠) as the header
        assert "&#x26A0;" in resp.data.decode("utf-8")


# ── Note field is always shown (not gated by visible_fields) ──────────────

class TestNoteAlwaysVisible:
    def test_note_field_always_in_form(self, client, seed_settings,
                                       db_session):
        seed_settings.enable_tire_details = False
        seed_settings.visible_fields = []
        db_session.commit()
        resp = client.get("/wheelsets/new")
        assert resp.status_code == 200
        assert b'name="note"' in resp.data

    def test_note_not_in_optional_fields(self):
        assert "note" not in Settings.OPTIONAL_FIELDS

    def test_note_not_in_settings_checkboxes(self, client, seed_settings):
        resp = client.get("/settings")
        assert resp.status_code == 200
        assert b"vf_note" not in resp.data
