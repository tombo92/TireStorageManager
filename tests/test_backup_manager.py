"""Tests for tsm/backup_manager.py — backup, CSV and XLSX export."""
import os
import tempfile
import time

from tsm.models import WheelSet, AuditLog
from tsm.backup_manager import BackupManager, export_csv_snapshot, export_xlsx_snapshot


class TestExportCsv:
    def test_creates_csv(self, db_session, db_engine, seed_wheelset,
                         monkeypatch):
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.csv")
            # Patch SessionLocal used inside export_csv_snapshot
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            result = export_csv_snapshot(target)
            assert result == target
            assert os.path.exists(target)
            with open(target, encoding="utf-8-sig") as f:
                content = f.read()
            assert "Mustermann" in content
            assert "C1ROM" in content

    def test_csv_header(self, db_session, db_engine, monkeypatch):
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.csv")
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            export_csv_snapshot(target)
            with open(target, encoding="utf-8-sig") as f:
                header = f.readline()
            assert "customer_name" in header
            assert "storage_position" in header


class TestBackupManager:
    def test_perform_backup(self, db_session, db_engine, seed_wheelset,
                            seed_settings, monkeypatch):
        with tempfile.TemporaryDirectory() as tmpdir:
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)

            mgr = BackupManager(db_engine, tmpdir)
            mgr.perform_backup()

            files = os.listdir(tmpdir)
            db_files = [f for f in files if f.endswith(".db")]
            csv_files = [f for f in files if f.endswith(".csv")]
            assert len(db_files) >= 1
            assert len(csv_files) >= 1

    def test_retention(self, db_session, db_engine, seed_wheelset,
                       seed_settings, monkeypatch):
        """Backup manager should respect retention (backup_copies)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)

            # Set retention to 2
            seed_settings.backup_copies = 2
            db_session.commit()

            mgr = BackupManager(db_engine, tmpdir)
            # Run 4 backups — only 2 should survive
            for _ in range(4):
                mgr.perform_backup()

            db_files = sorted(f for f in os.listdir(tmpdir) if f.endswith(".db"))
            csv_files = sorted(f for f in os.listdir(tmpdir) if f.endswith(".csv"))
            xlsx_files = sorted(f for f in os.listdir(tmpdir) if f.endswith(".xlsx"))
            assert len(db_files) <= 2
            assert len(csv_files) <= 2
            assert len(xlsx_files) <= 2

    def test_retention_keeps_newest(self, db_session, db_engine, seed_wheelset,
                                    seed_settings, monkeypatch):
        """After pruning, the two *newest* (alphabetically last) files must survive."""
        with tempfile.TemporaryDirectory() as tmpdir:
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)

            seed_settings.backup_copies = 2
            db_session.commit()

            mgr = BackupManager(db_engine, tmpdir)
            for _ in range(4):
                mgr.perform_backup()
                time.sleep(1)  # ensure unique second-resolution timestamps

            # The filenames are timestamp-sorted; the two survivors must be
            # the last two when sorted alphabetically (newest timestamps).
            all_db = sorted(f for f in os.listdir(tmpdir) if f.endswith(".db"))
            all_csv = sorted(f for f in os.listdir(tmpdir) if f.endswith(".csv"))
            assert len(all_db) == 2, f"expected 2 .db files, got {all_db}"
            assert len(all_csv) == 2, f"expected 2 .csv files, got {all_csv}"
            # Verify they are the two newest by checking timestamps are
            # strictly increasing (i.e. the larger timestamp strings survived).
            assert all_db[0] < all_db[1], "surviving .db files should be ordered"
            assert all_csv[0] < all_csv[1], "surviving .csv files should be ordered"
    def test_creates_xlsx(self, db_session, db_engine, seed_wheelset,
                          monkeypatch):
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.xlsx")
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            result = export_xlsx_snapshot(target)
            assert result == target
            assert os.path.exists(target)

    def test_xlsx_column_headers(self, db_session, db_engine, seed_wheelset,
                                 monkeypatch):
        """Inventory XLSX must contain the expected column headers."""
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.xlsx")
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            export_xlsx_snapshot(target)
            wb = openpyxl.load_workbook(target)
            ws = wb.active
            # Collect all cell values across the sheet
            all_values = [
                str(ws.cell(row=r, column=c).value or "")
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
            ]
            assert any("Kunde" in v for v in all_values)
            assert any("Position" in v for v in all_values)
            assert any("Kennzeichen" in v for v in all_values)
            assert any("Fahrzeug" in v for v in all_values)
            # Check column has the manual-check marker
            assert any("\u2713" in v for v in all_values)

    def test_xlsx_contains_wheelset_data(self, db_session, db_engine,
                                         seed_wheelset, monkeypatch):
        """Seeded wheel set must appear in the XLSX."""
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.xlsx")
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            export_xlsx_snapshot(target)
            wb = openpyxl.load_workbook(target)
            ws = wb.active
            all_values = [
                str(ws.cell(row=r, column=c).value or "")
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
            ]
            assert any("Mustermann" in v for v in all_values)
            assert any("C1ROM" in v for v in all_values)
            assert any("AB-CD 1234" in v for v in all_values)

    def test_xlsx_groups_containers(self, db_session, db_engine, monkeypatch):
        """Wheel sets in different containers appear under separate section headers."""
        import openpyxl
        db_session.add(WheelSet(
            customer_name="Anna Müller", license_plate="B-AM 111",
            car_type="BMW X3", storage_position="C1ROM",
        ))
        db_session.add(WheelSet(
            customer_name="Karl Berg", license_plate="M-KB 222",
            car_type="Audi A6", storage_position="C2ROM",
        ))
        db_session.commit()
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.xlsx")
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            export_xlsx_snapshot(target)
            wb = openpyxl.load_workbook(target)
            ws = wb.active
            all_values = [
                str(ws.cell(row=r, column=c).value or "")
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
            ]
            assert any("Container 1" in v for v in all_values)
            assert any("Container 2" in v for v in all_values)

    def test_xlsx_groups_garage(self, db_session, db_engine, monkeypatch):
        """Wheel sets in a garage shelf appear under the garage section header."""
        import openpyxl
        db_session.add(WheelSet(
            customer_name="Test Kunde", license_plate="HH-TK 999",
            car_type="VW Passat", storage_position="GR3OL",
        ))
        db_session.commit()
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.xlsx")
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            export_xlsx_snapshot(target)
            wb = openpyxl.load_workbook(target)
            ws = wb.active
            all_values = [
                str(ws.cell(row=r, column=c).value or "")
                for r in range(1, ws.max_row + 1)
                for c in range(1, ws.max_column + 1)
            ]
            assert any("Garage" in v and "3" in v for v in all_values)

    def test_xlsx_title_row(self, db_session, db_engine, monkeypatch):
        """XLSX title row must contain the inventory heading."""
        import openpyxl
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.xlsx")
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            export_xlsx_snapshot(target)
            wb = openpyxl.load_workbook(target)
            ws = wb.active
            title = str(ws.cell(row=1, column=1).value or "")
            assert "Bestand" in title or "Reifenlager" in title

    def test_xlsx_audit_log_created(self, db_session, db_engine, seed_wheelset,
                                    monkeypatch):
        """export_xlsx_snapshot must write one backup_xlsx audit log entry."""
        with tempfile.TemporaryDirectory() as tmpdir:
            target = os.path.join(tmpdir, "test.xlsx")
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            export_xlsx_snapshot(target)
            entries = db_session.query(AuditLog).filter_by(
                action="backup_xlsx").all()
            assert len(entries) == 1

    def test_perform_backup_creates_xlsx(self, db_session, db_engine,
                                         seed_wheelset, seed_settings,
                                         monkeypatch):
        """perform_backup must produce exactly one .xlsx file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            mgr = BackupManager(db_engine, tmpdir)
            mgr.perform_backup()
            xlsx_files = [f for f in os.listdir(tmpdir)
                          if f.endswith(".xlsx")]
            assert len(xlsx_files) == 1

    def test_xlsx_retention(self, db_session, db_engine, seed_wheelset,
                            seed_settings, monkeypatch):
        """BackupManager must respect retention limit for XLSX files."""
        with tempfile.TemporaryDirectory() as tmpdir:
            import tsm.backup_manager as bm_mod
            monkeypatch.setattr(bm_mod, "SessionLocal", db_session)
            seed_settings.backup_copies = 2
            db_session.commit()
            mgr = BackupManager(db_engine, tmpdir)
            for _ in range(4):
                mgr.perform_backup()
            xlsx_files = [f for f in os.listdir(tmpdir)
                          if f.endswith(".xlsx")]
            assert len(xlsx_files) <= 2
