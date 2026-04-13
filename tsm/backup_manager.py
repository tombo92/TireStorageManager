#!/usr/bin/env python
# @Date    : 2026-02-03 06:54:54
# @Author  : Tom Brandherm (https://github.com/tombo92)
# @Link    : https://github.com/tombo92/TireStorageManager
"""
Backup Manager
"""
# ========================================================
# IMPORTS
# ========================================================
import csv
import logging
import os
import sqlite3
import threading
from collections import defaultdict
from datetime import UTC, datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# --------------------------------------------------------
# Local Imports
# --------------------------------------------------------
from config import BACKUP_DIR
from tsm.db import SessionLocal
from tsm.models import AuditLog, Settings, WheelSet
from tsm.positions import RE_CONTAINER, RE_GARAGE, position_sort_key


# ========================================================
# CLASSES
# ========================================================
class BackupManager(threading.Thread):
    daemon = True

    _log = logging.getLogger("TSM.backup")

    def __init__(self, engine, backup_dir):
        super().__init__()
        self.engine = engine
        self.backup_dir = backup_dir
        self._stop_event = threading.Event()
        self._last_run = None

    def stop(self):
        self._stop_event.set()

    def run(self):
        while not self._stop_event.is_set():
            try:
                db = SessionLocal()
                try:
                    settings = db.query(Settings).first()
                    if settings is None:
                        settings = Settings(backup_interval_minutes=60,
                                            backup_copies=10)
                        db.add(settings)
                        db.commit()
                    interval = max(1, int(settings.backup_interval_minutes))
                    due = False
                    if self._last_run is None:
                        self._last_run = datetime.now(UTC)
                    elif ((datetime.now(UTC) - self._last_run)
                          >= timedelta(minutes=interval)):
                        due = True
                finally:
                    SessionLocal.remove()
                if due:
                    self.perform_backup()
                    self._last_run = datetime.now(UTC)
            except Exception:
                self._log.warning("BackupManager loop error",
                                  exc_info=True)
            self._stop_event.wait(30)

    def perform_backup(self):
        """Perform a backup of the database and export a CSV and XLSX snapshot. Old backups
        """
        ts = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
        bfile = os.path.join(self.backup_dir, f"wheel_storage_{ts}.db")

        raw = self.engine.raw_connection()
        try:
            src = raw.driver_connection  # sqlite3.Connection
            dest = sqlite3.connect(bfile)
            try:
                with dest:
                    src.backup(dest)
            finally:
                dest.close()
        finally:
            raw.close()

        csvfile = os.path.join(self.backup_dir, f"wheel_storage_{ts}.csv")
        export_csv_snapshot(csvfile)

        xlsxfile = os.path.join(self.backup_dir, f"wheel_storage_{ts}.xlsx")
        export_xlsx_snapshot(xlsxfile)

        db = SessionLocal()
        try:
            db.add(AuditLog(action="backup",
                            details=f"Backup erstellt: {os.path.basename(bfile)}"))
            db.commit()

            settings = db.query(Settings).first()
            keep = max(1, settings.backup_copies if settings else 10)

            backups_db = sorted(
                [f for f in os.listdir(self.backup_dir)
                 if f.startswith("wheel_storage_") and f.endswith(".db")]
            )
            if len(backups_db) > keep:
                for f in backups_db[0:len(backups_db)-keep]:
                    try:
                        os.remove(os.path.join(self.backup_dir, f))
                    except Exception:
                        pass

            backups_csv = sorted(
                [f for f in os.listdir(self.backup_dir)
                 if f.startswith("wheel_storage_") and f.endswith(".csv")]
            )
            if len(backups_csv) > keep:
                for f in backups_csv[0:len(backups_csv)-keep]:
                    try:
                        os.remove(os.path.join(self.backup_dir, f))
                    except Exception:
                        pass

            backups_xlsx = sorted(
                [f for f in os.listdir(self.backup_dir)
                 if f.startswith("wheel_storage_") and f.endswith(".xlsx")]
            )
            if len(backups_xlsx) > keep:
                for f in backups_xlsx[0:len(backups_xlsx)-keep]:
                    try:
                        os.remove(os.path.join(self.backup_dir, f))
                    except Exception:
                        pass
        finally:
            SessionLocal.remove()


# ========================================================
# FUNCTIONS
# ========================================================
def export_csv_snapshot(target_path: str | None = None) -> str:
    db = SessionLocal()
    try:
        rows = db.query(WheelSet).order_by(
            WheelSet.storage_position.asc()).all()
        if target_path is None:
            ts = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
            target_path = os.path.join(BACKUP_DIR, f"wheel_storage_{ts}.csv")
        with open(target_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=';')
            w.writerow(["customer_name", "license_plate", "car_type", "note",
                        "storage_position", "created_at", "updated_at"])
            for r in rows:
                w.writerow([
                    r.customer_name,
                    r.license_plate,
                    r.car_type,
                    r.note or "",
                    r.storage_position,
                    (r.created_at.isoformat() if r.created_at else ""),
                    (r.updated_at.isoformat() if r.updated_at else ""),
                ])
        filename = os.path.basename(target_path)
        db.add(AuditLog(action="backup_csv",
                        details=f"CSV exportiert: {filename}"))
        db.commit()
        return target_path
    finally:
        SessionLocal.remove()


def export_xlsx_snapshot(target_path: str | None = None) -> str:
    """Export a print-ready XLSX inventory grouped by container and garage."""
    db = SessionLocal()
    try:
        rows = db.query(WheelSet).all()
        rows = sorted(rows, key=lambda r: position_sort_key(r.storage_position))

        if target_path is None:
            ts = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
            target_path = os.path.join(BACKUP_DIR, f"wheel_storage_{ts}.xlsx")

        # Group by container (C1–C4) or garage shelf (GR1–GR8)
        groups_map: dict = defaultdict(list)
        for r in rows:
            m = RE_CONTAINER.match(r.storage_position)
            if m:
                key = ("container", int(m.group(1)), f"Container {m.group(1)}")
            else:
                m2 = RE_GARAGE.match(r.storage_position)
                if m2:
                    key = ("garage", int(m2.group(1)),
                           f"Garage \u2013 Regal {m2.group(1)}")
                else:
                    key = ("other", 0, "Sonstige Positionen")
            groups_map[key].append(r)

        sorted_groups = sorted(
            groups_map.items(),
            key=lambda x: (0 if x[0][0] == "container"
                           else 1 if x[0][0] == "garage" else 2, x[0][1])
        )

        # ── styles ────────────────────────────────────────────────────────
        CONTAINER_FILL = PatternFill("solid", fgColor="BDD7EE")
        GARAGE_FILL = PatternFill("solid", fgColor="C6EFCE")
        OTHER_FILL = PatternFill("solid", fgColor="FFFFCC")
        HEADER_FILL = PatternFill("solid", fgColor="2F75B6")
        ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
        THIN = Side(style="thin")
        FULL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        # ── workbook ──────────────────────────────────────────────────────
        wb = Workbook()
        ws = wb.active
        ws.title = "Bestandsübersicht"

        col_widths = [5, 14, 24, 14, 22, 24, 5]
        col_headers = ["Nr.", "Position", "Kunde", "Kennzeichen",
                       "Fahrzeug", "Notiz", "\u2713"]
        num_cols = len(col_headers)
        last_col = get_column_letter(num_cols)

        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Title
        ws.merge_cells(f"A1:{last_col}1")
        c = ws["A1"]
        c.value = "Reifenlager \u2013 Bestandsübersicht"
        c.font = Font(bold=True, size=14)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Sub-title with date and total count
        ws.merge_cells(f"A2:{last_col}2")
        c = ws["A2"]
        c.value = (
            f"Erstellt: "
            f"{datetime.now(UTC).strftime('%d.%m.%Y %H:%M')} UTC"
            f"  \u2013  Gesamt: {len(rows)} Rads\u00e4tze"
        )
        c.font = Font(italic=True, size=10)
        c.alignment = Alignment(horizontal="center")

        current_row = 3  # blank separator after header rows

        for (gtype, _gid, glabel), group_rows in sorted_groups:
            group_fill = (
                CONTAINER_FILL if gtype == "container"
                else GARAGE_FILL if gtype == "garage"
                else OTHER_FILL
            )

            # Blank row before each section (skip before first if immediately after header)
            current_row += 1

            # Section header
            ws.merge_cells(f"A{current_row}:{last_col}{current_row}")
            c = ws.cell(row=current_row, column=1,
                        value=f"  {glabel}  ({len(group_rows)} Rads\u00e4tze)")
            c.font = Font(bold=True, size=11)
            c.fill = group_fill
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    indent=1)
            ws.row_dimensions[current_row].height = 18

            # Column header row
            current_row += 1
            for col_idx, header in enumerate(col_headers, 1):
                c = ws.cell(row=current_row, column=col_idx, value=header)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = HEADER_FILL
                c.border = FULL_BORDER
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[current_row].height = 16

            # Data rows
            for i, r in enumerate(group_rows):
                current_row += 1
                fill = ALT_FILL if i % 2 == 1 else None
                values = [
                    i + 1,
                    r.storage_position,
                    r.customer_name,
                    r.license_plate,
                    r.car_type,
                    r.note or "",
                    "",
                ]
                for col_idx, val in enumerate(values, 1):
                    c = ws.cell(row=current_row, column=col_idx, value=val)
                    c.border = FULL_BORDER
                    c.alignment = Alignment(vertical="center")
                    if fill:
                        c.fill = fill
                    if col_idx == num_cols:
                        c.alignment = Alignment(horizontal="center",
                                                vertical="center")
                ws.row_dimensions[current_row].height = 15

        # Print settings: landscape, fit to one page wide
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToPage = True
        ws.page_margins.left = 0.5
        ws.page_margins.right = 0.5
        ws.page_margins.top = 0.75
        ws.page_margins.bottom = 0.75

        wb.save(target_path)
        db.add(AuditLog(
            action="backup_xlsx",
            details=f"XLSX exportiert: {os.path.basename(target_path)}"
        ))
        db.commit()
        return target_path
    finally:
        SessionLocal.remove()
